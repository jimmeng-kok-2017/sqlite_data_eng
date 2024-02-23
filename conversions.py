import openpyxl
import csv

def export_excel_work_book_sheet_to_csv(excel_file_name, sheet_name, csv_file_name):
    "open the excel file, find the sheet_name, export to csv file "
    
    work_book = openpyxl.load_workbook(excel_file_name)
    
    work_sheet = work_book[sheet_name]
                                       
    print("\nExcel Workbook:", excel_file_name)
          
    print("\nWork Sheet:", sheet_name)
    
    print("\nNumber of Columns:", work_sheet.max_column)
          
    print("\nNumber of Rows:", work_sheet.max_row)
    
    work_sheet_list = []
    
    for i in range(1,work_sheet.max_row+1):
        
        row_list = []
        
        for j in range(1,work_sheet.max_column+1):
            
            row_list.append(work_sheet.cell(row = i, column = j).value)
            
        work_sheet_list.append(row_list)
    
    print("\nExtracting Work Sheet", sheet_name, "to", csv_file_name, "\n")
    
    f = open(csv_file_name, "w")    
    writer = csv.writer(f)
    writer.writerows(work_sheet_list)
    f.close()
	
def read_csv_file(file_name, limit):
    "read the csv file and print only the first limit rows"
    
    csv_file = open(file_name, "r")
    
    csv_data = csv.reader(csv_file)
    
    i = 0
    
    for row in csv_data:
        i += 1
        if i <= limit:
            print(row)
            
    print("\nPrinted ", min(limit, i), "lines of ", i, "total lines.")
